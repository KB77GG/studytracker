"""Material Bank API for structured learning materials."""

from flask import Blueprint, request, jsonify
from flask_login import login_required, current_user
from models import db, MaterialBank, Question, QuestionOption, StudentAnswer, Task
from datetime import datetime
from openpyxl import load_workbook
import re

material_bp = Blueprint('material', __name__, url_prefix='/api/materials')
READING_VOCAB_CHOICE_TYPE = 'reading_vocab_choice'


def require_teacher():
    """Decorator to ensure user isadvertis teacher/admin."""
    if not current_user.is_authenticated:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    if current_user.role not in ['teacher', 'admin', 'assistant']:
        return jsonify({"ok": False, "error": "forbidden"}), 403
    return None


def _build_choice_questions(material_id, questions_data):
    for q_data in questions_data:
        question = Question(
            material_id=material_id,
            sequence=q_data.get('sequence'),
            question_type=q_data.get('question_type', 'choice'),
            content=q_data.get('content'),
            reference_answer=q_data.get('reference_answer'),
            hint=q_data.get('hint'),
            explanation=q_data.get('explanation', ''),
            points=q_data.get('points', 1)
        )
        db.session.add(question)
        db.session.flush()

        if q_data.get('options'):
            for opt in q_data['options']:
                option = QuestionOption(
                    question_id=question.id,
                    option_key=opt['key'],
                    option_text=opt['text']
                )
                db.session.add(option)


def _normalize_header(value):
    text = str(value or '').strip().lower()
    return text.replace(' ', '').replace('_', '')


def _parse_reading_vocab_choice_workbook(file_storage):
    try:
        wb = load_workbook(file_storage, data_only=True)
    except Exception as exc:
        raise ValueError(f"无法读取 Excel 文件: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel 为空")

    headers = [_normalize_header(v) for v in rows[0]]
    header_map = {name: idx for idx, name in enumerate(headers) if name}

    english_idx = None
    answer_idx = None
    correct_chinese_idx = None
    option_indices = {}

    for key, idx in header_map.items():
        if key in {'英文', 'english', 'word', '单词'}:
            english_idx = idx
        elif key in {'正确中文', 'correctchinese', 'correctmeaning', '正确释义'}:
            correct_chinese_idx = idx
        elif key in {'答案', 'answer'}:
            answer_idx = idx
        elif key in {'选项a', 'a', 'optiona'}:
            option_indices['A'] = idx
        elif key in {'选项b', 'b', 'optionb'}:
            option_indices['B'] = idx
        elif key in {'选项c', 'c', 'optionc'}:
            option_indices['C'] = idx
        elif key in {'选项d', 'd', 'optiond'}:
            option_indices['D'] = idx

    if english_idx is None:
        raise ValueError("缺少“英文”列")
    if answer_idx is None:
        raise ValueError("缺少“答案”列")
    if len(option_indices) != 4:
        raise ValueError("必须包含 4 个选项列：选项A/选项B/选项C/选项D")

    questions = []
    for row_no, row in enumerate(rows[1:], start=2):
        english = str(row[english_idx] or '').strip()
        if not english:
            continue

        options = []
        for key in ['A', 'B', 'C', 'D']:
            idx = option_indices[key]
            text = str(row[idx] or '').strip()
            if not text:
                raise ValueError(f"第 {row_no} 行缺少选项 {key}")
            options.append({'key': key, 'text': text})

        answer = str(row[answer_idx] or '').strip().upper()
        if answer not in {'A', 'B', 'C', 'D'}:
            raise ValueError(f"第 {row_no} 行答案必须是 A/B/C/D")

        correct_chinese = ''
        if correct_chinese_idx is not None:
            correct_chinese = str(row[correct_chinese_idx] or '').strip()
        if not correct_chinese:
            correct_chinese = next((opt['text'] for opt in options if opt['key'] == answer), '')

        questions.append({
            'sequence': len(questions) + 1,
            'question_type': 'choice',
            'content': english,
            'reference_answer': answer,
            'hint': correct_chinese,
            'explanation': '',
            'points': 1,
            'options': options,
        })

    if not questions:
        raise ValueError("未识别到有效题目")

    return questions


# ============================================================================
# Material CRUD
# ============================================================================

@material_bp.route('', methods=['GET'])
@login_required
def list_materials():
    """Get list of materials with optional filtering."""
    auth_check = require_teacher()
    if auth_check:
        return auth_check
    
    material_type = request.args.get('type')
    search = request.args.get('search', '').strip()
    
    query = MaterialBank.query.filter_by(is_deleted=False, is_active=True)
    
    if material_type:
        query = query.filter_by(type=material_type)
    
    if search:
        query = query.filter(MaterialBank.title.contains(search))
    
    query = query.order_by(MaterialBank.created_at.desc())
    materials = query.all()
    
    result = []
    for m in materials:
        question_count = Question.query.filter_by(material_id=m.id).count()
        result.append({
            "id": m.id,
            "title": m.title,
            "type": m.type,
            "description": m.description,
            "question_count": question_count,
            "created_at": m.created_at.strftime("%Y-%m-%d")
        })
    
    return jsonify({"ok": True, "materials": result})


@material_bp.route('/<int:material_id>', methods=['GET'])
@login_required
def get_material(material_id):
    """Get single material with questions."""
    auth_check = require_teacher()
    if auth_check:
        return auth_check
    
    material = MaterialBank.query.filter_by(id=material_id, is_deleted=False).first()
    if not material:
        return jsonify({"ok": False, "error": "material_not_found"}), 404
    
    questions = Question.query.filter_by(material_id=material_id).order_by(Question.sequence).all()
    
    questions_data = []
    for q in questions:
        options = QuestionOption.query.filter_by(question_id=q.id).order_by(QuestionOption.option_key).all()
        questions_data.append({
            "id": q.id,
            "sequence": q.sequence,
            "question_type": q.question_type,
            "content": q.content,
            "reference_answer": q.reference_answer,
            "hint": q.hint,
            "points": q.points,
            "options": [{"key": opt.option_key, "text": opt.option_text} for opt in options]
        })
    
    return jsonify({
        "ok": True,
        "material": {
            "id": material.id,
            "title": material.title,
            "type": material.type,
            "description": material.description,
            "questions": questions_data
        }
    })


@material_bp.route('', methods=['POST'])
@login_required
def create_material():
    """Create new material with questions."""
    auth_check = require_teacher()
    if auth_check:
        return auth_check
    
    data = request.get_json()
    material_type = data.get('type')
    
    # Create material
    material = MaterialBank(
        title=data.get('title'),
        type=material_type,
        description=data.get('description', ''),
        created_by=current_user.id
    )
    db.session.add(material)
    db.session.flush()  # Get material.id
    
    # Handle different material types
    if material_type in {'grammar', READING_VOCAB_CHOICE_TYPE}:
        # Grammar: traditional questions with options
        questions_data = data.get('questions', [])
        _build_choice_questions(material.id, questions_data)
    
    elif material_type == 'writing_logic':
        # Writing logic chain: single question with essay data
        question = Question(
            material_id=material.id,
            sequence=1,
            question_type='writing_logic',
            content=data.get('essay_full'),  # Full essay
            hint=data.get('essay_blank'),  # Blank version with labels
            reference_answer=data.get('essay_answers')  # Answer list
        )
        db.session.add(question)
    
    elif material_type == 'speaking_reading':
        # Speaking reading: single question with vocabulary, patterns, and paragraph
        question = Question(
            material_id=material.id,
            sequence=1,
            question_type='speaking_reading',
            content=data.get('vocabulary', ''),  # Vocabulary expressions
            hint=data.get('sentence_patterns', ''),  # Sentence patterns
            reference_answer=data.get('sample_paragraph', '')  # Sample paragraph
        )
        db.session.add(question)
    
    elif material_type == 'speaking_part1':
        # Speaking Part 1: multiple questions, each as a separate record
        questions_text = data.get('part1_questions', '')
        questions_list = _parse_speaking_part1(questions_text)
        if not questions_list:
            questions_list = [q.strip() for q in questions_text.strip().split('\n') if q.strip()]
        
        for i, question_text in enumerate(questions_list, 1):
            question = Question(
                material_id=material.id,
                sequence=i,
                question_type='speaking_part1',
                content=question_text
            )
            db.session.add(question)
    
    elif material_type == 'speaking_part2':
        # Speaking Part 2: single question with topic card
        question = Question(
            material_id=material.id,
            sequence=1,
            question_type='speaking_part2',
            content=data.get('part2_topic')  # Topic card text
        )
        db.session.add(question)

    elif material_type == 'speaking_part2_3':
        # Speaking Part 2+3: multiple topic cards, each stored as one question
        cards_text = data.get('part23_topics', '')
        cards = _parse_speaking_part23(cards_text)
        for i, card in enumerate(cards, 1):
            question = Question(
                material_id=material.id,
                sequence=i,
                question_type='speaking_part2_3',
                content=card,
            )
            db.session.add(question)
    
    elif material_type == 'translation':
        # Translation exercise: single question with sentence, skeleton, grammar, and reference
        question = Question(
            material_id=material.id,
            sequence=1,
            question_type='translation',
            content=data.get('sentence', ''),  # Original English sentence
            hint=data.get('skeleton', ''),  # Sentence skeleton structure
            explanation=data.get('grammar', ''),  # Grammar points
            reference_answer=data.get('reference', '')  # Reference translation
        )
        db.session.add(question)
    
    db.session.commit()
    
    return jsonify({"ok": True, "material_id": material.id})


@material_bp.route('/<int:material_id>', methods=['PUT'])
@login_required
def update_material(material_id):
    """Update material."""
    auth_check = require_teacher()
    if auth_check:
        return auth_check
    
    material = MaterialBank.query.filter_by(id=material_id, is_deleted=False).first()
    if not material:
        return jsonify({"ok": False, "error": "material_not_found"}), 404
    
    data = request.get_json()
    
    # Update basic info
    if 'title' in data:
        material.title = data['title']
    if 'type' in data:
        material.type = data['type']
    if 'description' in data:
        material.description = data['description']
    
    # If questions are provided, replace all questions
    if 'questions' in data:
        # Delete existing questions (cascade will delete options)
        Question.query.filter_by(material_id=material_id).delete()
        
        # Create new questions
        questions_data = data.get('questions', [])
        _build_choice_questions(material.id, questions_data)
    elif material.type == 'speaking_part1' and data.get('part1_questions'):
        Question.query.filter_by(material_id=material_id).delete()
        questions_list = _parse_speaking_part1(data.get('part1_questions', ''))
        if not questions_list:
            questions_list = [q.strip() for q in data.get('part1_questions', '').split('\n') if q.strip()]
        for i, question_text in enumerate(questions_list, 1):
            question = Question(
                material_id=material.id,
                sequence=i,
                question_type='speaking_part1',
                content=question_text
            )
            db.session.add(question)
    elif material.type == 'speaking_part2' and data.get('part2_topic'):
        Question.query.filter_by(material_id=material_id).delete()
        question = Question(
            material_id=material.id,
            sequence=1,
            question_type='speaking_part2',
            content=data.get('part2_topic')
        )
        db.session.add(question)
    elif material.type == 'speaking_part2_3' and data.get('part23_topics'):
        Question.query.filter_by(material_id=material_id).delete()
        cards = _parse_speaking_part23(data.get('part23_topics', ''))
        for i, card in enumerate(cards, 1):
            question = Question(
                material_id=material.id,
                sequence=i,
                question_type='speaking_part2_3',
                content=card,
            )
            db.session.add(question)
    
    material.updated_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({"ok": True})


@material_bp.route('/<int:material_id>', methods=['DELETE'])
@login_required
def delete_material(material_id):
    """Soft delete material."""
    auth_check = require_teacher()
    if auth_check:
        return auth_check
    
    material = MaterialBank.query.filter_by(id=material_id, is_deleted=False).first()
    if not material:
        return jsonify({"ok": False, "error": "material_not_found"}), 404
    
    material.is_deleted = True
    material.updated_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({"ok": True})


@material_bp.route('/parse-reading-vocab-choice', methods=['POST'])
@login_required
def parse_reading_vocab_choice():
    """Parse an Excel workbook of English-to-Chinese vocab choice questions."""
    auth_check = require_teacher()
    if auth_check:
        return auth_check

    if "file" not in request.files:
        return jsonify({"ok": False, "error": "missing_file"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"ok": False, "error": "empty_filename"}), 400

    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({"ok": False, "error": "invalid_extension", "message": "仅支持 .xlsx 文件"}), 400

    try:
        questions = _parse_reading_vocab_choice_workbook(file)
        return jsonify({"ok": True, "questions": questions})
    except ValueError as exc:
        return jsonify({"ok": False, "error": "parse_failed", "message": str(exc)}), 400


# ============================================================================
# Question Parsing Utilities
# ============================================================================

def _parse_speaking_part1(text: str) -> list[str]:
    lines = [ln.strip() for ln in (text or '').split('\n')]
    questions: list[str] = []
    current_topic = ''
    has_topic = False

    for line in lines:
        if not line:
            continue

        header_match = re.match(r'^(\d+)[.、)]\s*(.+)$', line)
        if header_match:
            current_topic = header_match.group(2).strip()
            has_topic = True
            continue

        bullet_match = re.match(r'^[*•\-]\s*(.+)$', line)
        if bullet_match:
            question_text = bullet_match.group(1).strip()
        else:
            question_text = line

        if not question_text:
            continue

        if has_topic and current_topic:
            question_text = f"{current_topic} - {question_text}"
        questions.append(question_text)

    return questions


def _looks_like_part23_title(line: str) -> bool:
    text = (line or "").strip()
    if not text:
        return False
    lower = text.lower()

    # Not a card title
    if text.startswith(('*', '•', '-', '·')):
        return False
    if re.match(r"^part\s*3\b", lower):
        return False
    if re.match(r"^you should say[:：]?$", lower):
        return False
    if re.match(r"^describe\b", lower):
        return False
    if re.search(r"[?？!！。]$", text):
        return False

    # Typical title patterns
    if re.match(r"^\d+[.)、]\s*", text):
        return True
    if re.search(r"[（(].+[)）]", text):
        return True
    if re.search(r"[\u4e00-\u9fff]", text) and len(text) <= 40:
        return True
    if re.match(r"^[A-Z][A-Za-z0-9/&,'\-\s]{2,40}$", text):
        return True
    return False


def _parse_speaking_part23(text: str) -> list[str]:
    raw = (text or "").replace("\r\n", "\n").strip()
    if not raw:
        return []

    lines = raw.split("\n")
    cards: list[str] = []
    current: list[str] = []

    for line in lines:
        stripped = line.strip()
        if not stripped:
            if current and current[-1] != "":
                current.append("")
            continue

        if _looks_like_part23_title(stripped) and current:
            block = "\n".join(current).strip()
            if block:
                cards.append(block)
            current = []

        current.append(stripped)

    if current:
        block = "\n".join(current).strip()
        if block:
            cards.append(block)

    if cards:
        return cards

    blocks = re.split(r"\n\s*\n+", raw)
    return [block.strip() for block in blocks if block.strip()]

@material_bp.route('/parse', methods=['POST'])
@login_required
def parse_questions():
    """Parse questions from pasted text (Word content)."""
    auth_check = require_teacher()
    if auth_check:
        return auth_check
    
    data = request.get_json()
    text = data.get('text', '')
    
    questions = []
    lines = text.split('\n')
    
    current_question = None
    current_options = []
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Match question number: "1. " or "1、"
        question_match = re.match(r'^(\d+)[.、]\s*(.+)$', line)
        if question_match:
            # Save previous question
            if current_question:
                current_question['options'] = current_options
                questions.append(current_question)
            
            # Start new question
            sequence = int(question_match.group(1))
            content = question_match.group(2)
            current_question = {
                'sequence': sequence,
                'content': content,
                'question_type': 'choice',
                'reference_answer': '',
                'points': 1
            }
            current_options = []
            continue
        
        # Match options: "A. " or "A、"
        option_match = re.match(r'^([A-D])[.、]\s*(.+)$', line)
        if option_match and current_question:
            option_key = option_match.group(1)
            option_text = option_match.group(2)
            current_options.append({'key': option_key, 'text': option_text})
            continue
        
        # Otherwise, append to current question content
        if current_question and not option_match:
            current_question['content'] += ' ' + line
    
    # Don't forget last question
    if current_question:
        current_question['options'] = current_options
        questions.append(current_question)
    
    return jsonify({"ok": True, "questions": questions})


@material_bp.route('/parse-answers', methods=['POST'])
@login_required
def parse_answers():
    """Parse answers from pasted text."""
    auth_check = require_teacher()
    if auth_check:
        return auth_check
    
    data = request.get_json()
    text = data.get('text', '')
    
    answers = {}
    
    # Normalize text: replace Chinese punctuation with English
    text = text.replace('：', ':').replace('、', '.')
    
    # Strategy 1: Look for multiple choice patterns (multiple on one line)
    # e.g. "1. A 2. B" or "7. B/C"
    # Regex explanation:
    # (\d+)       : Capture Group 1 - Question number
    # \s*[.]\s*   : Dot separator with optional whitespace
    # (?:Answer:)? : Optional "Answer:" prefix
    # \s*         : Optional whitespace
    # ([A-E]+(?:/[A-E]+)*) : Capture Group 2 - Answer (e.g. "A", "AB", "B/C")
    choice_matches = re.findall(r'(\d+)\s*[.]\s*(?:Answer:)?\s*([A-E]+(?:/[A-E]+)*)', text, re.IGNORECASE)
    
    for num_str, ans_str in choice_matches:
        answers[int(num_str)] = ans_str.upper()
    
    # Strategy 2: Look for text answers (one per line)
    # This handles "1. The government built..."
    lines = text.split('\n')
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Check if line looks like it contains multiple choice answers (handled by Strategy 1)
        # e.g. "1. A 2. B" - we skip these lines to avoid overwriting with partial text
        if re.search(r'\d+\s*[.]\s*[A-E]\s+\d+', line):
            continue
            
        # Match "Number. Content"
        match = re.match(r'^(\d+)\s*[.]\s*(.+)$', line)
        if match:
            seq = int(match.group(1))
            content = match.group(2).strip()
            
            # Only use this if we haven't found a choice answer for this sequence
            # OR if the content is clearly longer than a choice answer (e.g. > 5 chars)
            # This prevents "1. A" (text match) from overwriting "1. A" (choice match), which is fine,
            # but helps if Strategy 1 missed something.
            if seq not in answers or len(content) > 5:
                # Special case: if content is just "A" or "B", treat as choice
                if re.match(r'^[A-E]$', content, re.IGNORECASE):
                    answers[seq] = content.upper()
                else:
                    answers[seq] = content
    
    return jsonify({"ok": True, "answers": answers})


@material_bp.route('/parse-pdf', methods=['POST'])
@login_required
def parse_pdf():
    """Parse grammar questions from uploaded PDF or PPTX file."""
    auth_check = require_teacher()
    if auth_check:
        return auth_check

    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "no_file"}), 400

    file = request.files['file']
    filename = file.filename.lower()

    if filename.endswith('.pptx'):
        return _parse_pptx_upload(file)
    elif filename.endswith('.pdf'):
        return _parse_pdf_upload(file)
    else:
        return jsonify({"ok": False, "error": "unsupported_format", "message": "请上传 PDF 或 PPTX 文件"}), 400


def _parse_pdf_upload(file):
    """Parse grammar questions from PDF file."""
    import pdfplumber
    import tempfile
    import os

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    file.save(tmp.name)
    tmp.close()

    try:
        with pdfplumber.open(tmp.name) as pdf:
            all_text = []
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text.append(text)
            full_text = '\n'.join(all_text)
    finally:
        os.unlink(tmp.name)

    questions = _parse_pdf_questions(full_text)
    return jsonify({"ok": True, "questions": questions, "raw_text_length": len(full_text)})


def _parse_pptx_upload(file):
    """Parse grammar questions from PPTX file. Much better than PDF for preserving blanks."""
    import zipfile
    import tempfile
    import os
    from lxml import etree

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pptx')
    file.save(tmp.name)
    tmp.close()

    questions = []
    try:
        with zipfile.ZipFile(tmp.name) as z:
            slide_files = sorted(
                [f for f in z.namelist() if f.startswith('ppt/slides/slide') and f.endswith('.xml')],
                key=lambda x: int(re.search(r'(\d+)', x.split('/')[-1]).group(1))
            )
            ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}

            for sf in slide_files:
                with z.open(sf) as f:
                    tree = etree.parse(f)

                # Extract all paragraphs from the slide
                paragraphs = tree.findall('.//a:p', ns)
                slide_text = []
                for p in paragraphs:
                    runs = p.findall('.//a:t', ns)
                    para_text = ''.join(r.text for r in runs if r.text)
                    if para_text.strip():
                        slide_text.append(para_text.strip())

                full_slide = '\n'.join(slide_text)
                if not full_slide.strip():
                    continue

                # Try to parse question from this slide
                q = _parse_pptx_slide(full_slide)
                if q:
                    questions.append(q)
    finally:
        os.unlink(tmp.name)

    # Re-number sequences
    for i, q in enumerate(questions):
        q['sequence'] = i + 1

    return jsonify({"ok": True, "questions": questions, "source": "pptx"})


def _parse_pptx_slide(text):
    """Parse a single slide's text into a question dict.

    PPT slides typically have format:
      N.(source) stem with 　　 blank.
      A.opt1　　B.opt2　　C.opt3　　D.opt4
      答案    X　explanation...
    """
    # Normalize whitespace: full-width spaces → blank marker
    # PPT uses multiple full-width spaces (　) or regular spaces for blanks
    # Replace sequences of full-width spaces (with optional regular spaces) with ______
    text_normalized = re.sub(r'[　\u3000]{2,}[\s]*', ' ______ ', text)
    text_normalized = re.sub(r'\s*______\s*', ' ______ ', text_normalized)

    # Must have a question number pattern
    q_match = re.match(r'(\d+)\s*[.、．]\s*(?:\(([^)]*)\))?\s*(.*)', text_normalized, re.DOTALL)
    if not q_match:
        return None

    seq = int(q_match.group(1))
    source = q_match.group(2) or ''
    rest = q_match.group(3).strip()

    # Find answer section
    ans_match = re.search(r'答案\s*([A-D])\s*', rest)
    if not ans_match:
        return None

    answer = ans_match.group(1)
    q_part = rest[:ans_match.start()].strip()

    # Extract explanation (everything after "答案 X" up to 知识拓展/易错警示)
    exp_text = rest[ans_match.end():]
    exp_text = re.split(r'知识拓展|易错警示|方法技巧|思路分析', exp_text)[0]
    explanation = exp_text.strip()
    # Clean up explanation - remove leading punctuation
    explanation = re.sub(r'^[　\s]+', '', explanation)

    # Split stem and options
    # Options pattern: A.xxx B.xxx or A.xxx\nB.xxx
    # Find where options start
    lines = q_part.split('\n')
    stem_lines = []
    option_lines = []
    found_options = False

    for line in lines:
        line = line.strip()
        if not line:
            continue
        if re.match(r'^[A-D][.、．]', line) or (not found_options and re.search(r'\bA[.、．]', line)):
            option_lines.append(line)
            found_options = True
        elif found_options and re.match(r'^[B-D][.、．]', line):
            option_lines.append(line)
        else:
            if not found_options:
                stem_lines.append(line)

    stem = ' '.join(stem_lines).strip()
    # Clean extra spaces but preserve ______
    stem = re.sub(r'\s{2,}', ' ', stem)

    # Parse options - PPT often uses 　　 to separate options on one line
    opt_text = ' '.join(option_lines)
    # Normalize option separators
    opt_text = re.sub(r'[　\u3000]+', '  ', opt_text)

    options = []
    # Try matching options
    opt_matches = re.findall(r'([A-D])[.、．]\s*(.+?)(?=\s{2,}[B-D][.、．]|$)', opt_text)
    for key, text in opt_matches:
        options.append({'key': key, 'text': text.strip()})

    if len(options) < 2:
        # Fallback: split by option letter
        options = []
        parts = re.split(r'(?=[A-D][.、．])', opt_text)
        for part in parts:
            m = re.match(r'([A-D])[.、．]\s*(.+)', part.strip())
            if m:
                options.append({'key': m.group(1), 'text': m.group(2).strip()})

    if not options or not stem:
        return None

    # Ensure blank marker exists
    if '______' not in stem and '___' not in stem:
        stem = _ensure_blank_marker(stem, options, answer)

    hint = f"来源: {source}" if source else ''

    return {
        'sequence': seq,
        'content': stem,
        'question_type': 'choice',
        'reference_answer': answer,
        'explanation': explanation,
        'hint': hint,
        'points': 1,
        'options': options
    }


def _ensure_blank_marker(stem, options, answer):
    """Ensure the question stem contains a ______ blank marker.

    PDF underlines are often graphic elements lost during text extraction.
    This function tries all option texts at every word boundary to find the
    position where inserting an option best reconstructs the original sentence.
    """
    if '___' in stem or '______' in stem:
        return stem  # Already has blank marker

    # Collect all option texts (any could be the answer)
    opt_texts = [opt['text'] for opt in options if opt.get('text')]
    if not opt_texts:
        return stem

    # For each option, try inserting it at every word boundary.
    # Score by checking if the inserted text creates bigrams that exist
    # elsewhere in common English (simple: check if left+answer or answer+right
    # forms a known pattern). But simpler: the correct position is where
    # the option text is NOT already present as a substring.

    # Use the longest option text for better matching
    opt_texts_sorted = sorted(opt_texts, key=len, reverse=True)
    words = stem.split(' ')

    # Skip if a multi-word option text is already in stem (blank is elsewhere)
    # Don't skip for single-word options (articles, short words often appear naturally)
    for ot in opt_texts:
        if len(ot.split()) > 1 and ot.lower() in stem.lower():
            return stem

    best_pos = -1
    best_score = -1

    for opt_t in opt_texts_sorted[:2]:  # Try top 2 longest options
        opt_words = opt_t.split()
        for i in range(1, len(words)):  # Skip position 0
            # Build the sentence with option inserted
            left_word = words[i - 1].rstrip('.,;:!?"\u2019\u201d').lower()
            right_word = words[i].lstrip('.,;:!?"\u2018\u201c').lower() if i < len(words) else ''
            first_opt = opt_words[0].lower()
            last_opt = opt_words[-1].lower()

            score = 0

            # Subject + verb pattern (most common blank type)
            subjects = {'i', 'he', 'she', 'we', 'they', 'it', 'you', 'who', 'that',
                        'which', 'people', 'students', 'children', 'animals', 'doctors',
                        'teachers', 'scientists', 'everyone', 'someone', 'nobody'}
            if left_word in subjects:
                score += 3

            # After dialogue dash pattern: —...subject verb
            if left_word in subjects and i >= 2:
                prev2 = words[i-2] if i >= 2 else ''
                if prev2.startswith('—') or prev2.startswith('-'):
                    score += 2

            # Verb + preposition/object pattern
            preps = {'to', 'the', 'a', 'an', 'in', 'on', 'at', 'for', 'by', 'with',
                     'from', 'up', 'out', 'as', 'about', 'into', 'over'}
            if right_word in preps:
                score += 1

            # After auxiliary/modal
            auxs = {'will', 'would', 'shall', 'should', 'can', 'could', 'may',
                    'might', 'must', 'have', 'has', 'had', 'do', 'does', 'did',
                    "don't", "doesn't", "didn't", "won't", "can't", "haven't",
                    "hasn't", "hadn't", 'not', 'never', 'already', 'just', 'also'}
            if left_word in auxs:
                score += 3

            # After 'be' verbs (passive/continuous)
            if left_word in ('is', 'are', 'was', 'were', 'be', 'been', 'being'):
                score += 2

            # Article blanks: check if options are articles (the/a/an/∅)
            is_article_q = all(o.lower().strip() in ('the', 'a', 'an', '/', '∅', '') for o in opt_texts if o.strip())
            if is_article_q:
                # Articles go before nouns/adjectives (capitalized words or common patterns)
                # Look for position where an article naturally fits: before adjective/noun
                if right_word and right_word[0].isalpha() and left_word in (
                    'in', 'of', 'on', 'at', 'for', 'by', 'with', 'from', 'to',
                    'is', 'are', 'was', 'were', 'lies'):
                    score += 3
                # Beginning of sentence (after period or dash)
                if i == 1 or (i >= 1 and words[i-1].endswith(('.', '!', '?'))):
                    score += 2

            # Penalty: left word is a preposition (unlikely blank after prep for verb questions)
            if left_word in preps and score == 0 and not is_article_q:
                score -= 1

            if score > best_score:
                best_score = score
                best_pos = i

    if best_pos >= 0 and best_score >= 2:
        words.insert(best_pos, '______')
        return ' '.join(words)

    return stem


def _parse_pdf_questions(text):
    """Parse grammar questions from PDF text.

    Expected format per question:
      N.(source info)Question stem with ______ blank.
      A.option1  B.option2
      C.option3  D.option4
      答案 X  explanation text...
      [Optional: 思路分析 / 方法技巧 sections]
    """
    # Clean up duplicated header artifacts like 栏栏目目引索引
    text = re.sub(r'栏+目+引?索?引?', '', text)

    # Split into question blocks using the numbered pattern
    # Pattern: digit(s) followed by .( which indicates start of a question
    q_starts = list(re.finditer(r'(?:^|\n)\s*(\d+)\s*\.\s*\(', text))

    questions = []
    for i, match in enumerate(q_starts):
        start = match.start()
        end = q_starts[i + 1].start() if i + 1 < len(q_starts) else len(text)
        block = text[start:end].strip()

        q = _parse_single_question(block, i + 1)
        if q:
            questions.append(q)

    # If no questions found with source pattern, try simpler numbering
    if not questions:
        q_starts = list(re.finditer(r'(?:^|\n)\s*(\d+)\s*[.、]\s*', text))
        for i, match in enumerate(q_starts):
            start = match.start()
            end = q_starts[i + 1].start() if i + 1 < len(q_starts) else len(text)
            block = text[start:end].strip()
            q = _parse_single_question_simple(block, int(match.group(1)))
            if q:
                questions.append(q)

    return questions


def _parse_single_question(block, fallback_seq):
    """Parse a single question block with source info pattern like N.(2019江西,30)."""
    # Extract question number and source
    header = re.match(r'(\d+)\s*\.\s*\(([^)]+)\)\s*', block)
    if not header:
        return None

    seq = int(header.group(1))
    source = header.group(2).strip()
    rest = block[header.end():]

    # Find answer section
    ans_match = re.search(r'答案\s+([A-D])\s+', rest)
    if not ans_match:
        # Try without space after letter
        ans_match = re.search(r'答案\s+([A-D])', rest)

    answer = ans_match.group(1) if ans_match else ''
    explanation = ''

    if ans_match:
        # Everything after "答案 X" up to 思路分析/方法技巧 is explanation
        exp_text = rest[ans_match.end():]
        # Remove 思路分析 and 方法技巧 sections
        exp_text = re.split(r'思路分析|方法技巧', exp_text)[0]
        explanation = exp_text.strip()
        # The question+options part is before the answer
        q_part = rest[:ans_match.start()].strip()
    else:
        q_part = rest.strip()

    # Parse options from q_part
    # Options can be on same line: "A.xxx B.xxx" or separate lines: "A.xxx\nB.xxx"
    # Find all options
    opt_pattern = r'([A-D])\.\s*([^\n]+?)(?=\s+[B-D]\.|$|\n)'
    options = []

    # Try to find the options area (last lines of q_part that contain A. B. C. D.)
    lines = q_part.split('\n')
    stem_lines = []
    option_lines = []
    found_options = False

    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Check if line starts with or contains A. pattern
        if re.match(r'^[A-D]\.', line) or (not found_options and re.search(r'\bA\.', line)):
            option_lines.append(line)
            found_options = True
        elif found_options and re.match(r'^[A-D]\.', line):
            option_lines.append(line)
        else:
            if not found_options:
                stem_lines.append(line)
            else:
                # Could be continuation of options on new line
                if re.search(r'^[A-D]\.', line):
                    option_lines.append(line)
                else:
                    stem_lines.append(line)

    stem = ' '.join(stem_lines).strip()
    # Clean up stem - remove multiple spaces
    stem = re.sub(r'\s{2,}', ' ', stem)

    # Parse individual options from option_lines
    opt_text = ' '.join(option_lines)
    # Split by option letter pattern
    opt_matches = re.findall(r'([A-D])\.\s*(\S+(?:\s+\S+)*?)(?=\s+[B-D]\.\s*\S|$)', opt_text)

    for key, text in opt_matches:
        options.append({'key': key, 'text': text.strip()})

    # If we didn't get options, try a different approach
    if len(options) < 2:
        options = []
        for line in option_lines:
            single_opts = re.findall(r'([A-D])\.\s*(.+?)(?=\s{2,}[A-D]\.|$)', line)
            if single_opts:
                for key, text in single_opts:
                    options.append({'key': key, 'text': text.strip()})
            else:
                m = re.match(r'([A-D])\.\s*(.+)', line)
                if m:
                    options.append({'key': m.group(1), 'text': m.group(2).strip()})

    # Add blank marker if missing (PDF graphics-based underlines are lost during text extraction)
    stem = _ensure_blank_marker(stem, options, answer)

    # Add source to hint
    hint = f"来源: {source}"

    return {
        'sequence': seq,
        'content': stem,
        'question_type': 'choice',
        'reference_answer': answer,
        'explanation': explanation,
        'hint': hint,
        'points': 1,
        'options': options
    }


def _parse_single_question_simple(block, seq):
    """Parse a simple question block without source info."""
    # Find answer section
    ans_match = re.search(r'答案\s+([A-D])', block)
    answer = ans_match.group(1) if ans_match else ''
    explanation = ''

    if ans_match:
        exp_text = block[ans_match.end():]
        exp_text = re.split(r'思路分析|方法技巧', exp_text)[0]
        explanation = exp_text.strip()
        q_part = block[:ans_match.start()].strip()
    else:
        q_part = block.strip()

    # Remove leading number
    q_part = re.sub(r'^\d+\s*[.、]\s*', '', q_part)

    lines = q_part.split('\n')
    stem_lines = []
    option_lines = []

    for line in lines:
        line = line.strip()
        if not line:
            continue
        if re.match(r'^[A-D]\.', line) or re.search(r'\bA\.', line):
            option_lines.append(line)
        elif option_lines and re.match(r'^[C-D]\.', line):
            option_lines.append(line)
        else:
            if not option_lines:
                stem_lines.append(line)

    stem = ' '.join(stem_lines).strip()
    stem = re.sub(r'\s{2,}', ' ', stem)

    opt_text = ' '.join(option_lines)
    options = []
    opt_matches = re.findall(r'([A-D])\.\s*(\S+(?:\s+\S+)*?)(?=\s+[B-D]\.\s*\S|$)', opt_text)
    for key, text in opt_matches:
        options.append({'key': key, 'text': text.strip()})

    if len(options) < 2:
        options = []
        for line in option_lines:
            m = re.match(r'([A-D])\.\s*(.+)', line)
            if m:
                options.append({'key': m.group(1), 'text': m.group(2).strip()})

    # Add blank marker if missing
    stem = _ensure_blank_marker(stem, options, answer)

    return {
        'sequence': seq,
        'content': stem,
        'question_type': 'choice',
        'reference_answer': answer,
        'explanation': explanation,
        'hint': '',
        'points': 1,
        'options': options
    } if stem else None
